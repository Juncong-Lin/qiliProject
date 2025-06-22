import importlib.util
import sys
import os
import re
import requests
import time
import random
import pandas as pd
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Font, Alignment

# Dependency Check
required_modules = {
    'requests': 'requests',
    'bs4': 'beautifulsoup4',
    'pandas': 'pandas',
    'openpyxl': 'openpyxl'
}
for module, pip_name in required_modules.items():
    if importlib.util.find_spec(module) is None:
        print(f"Error: '{pip_name}' is not installed. Please run 'pip install {pip_name}' to install it.")
        sys.exit(1)

# Configuration
BASE_URL = "https://www.sign-in-global.com"
URL_TEMPLATE = "https://www.sign-in-global.com/product/index.php?page={page_num}&act=shopping_history&productid=20344"
TOTAL_PAGES = 260
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}
MAX_RETRIES = 6

# Comprehensive Country to Continent Mapping
COUNTRY_CONTINENT = {
    # Africa
    'Algeria': 'Africa', 'Angola': 'Africa', 'Benin': 'Africa', 'Botswana': 'Africa', 'Burkina Faso': 'Africa',
    'Burundi': 'Africa', 'Cameroon': 'Africa', 'Cape Verde': 'Africa', 'Central African Republic': 'Africa',
    'Chad': 'Africa', 'Comoros': 'Africa', 'Congo': 'Africa', 'Democratic Republic of the Congo': 'Africa',
    'Djibouti': 'Africa', 'Egypt': 'Africa', 'Equatorial Guinea': 'Africa', 'Eritrea': 'Africa', 'Eswatini': 'Africa',
    'Ethiopia': 'Africa', 'Gabon': 'Africa', 'Gambia': 'Africa', 'Ghana': 'Africa', 'Guinea': 'Africa',
    'Guinea-Bissau': 'Africa', 'Ivory Coast': 'Africa', 'Kenya': 'Africa', 'Lesotho': 'Africa', 'Liberia': 'Africa',
    'Libya': 'Africa', 'Madagascar': 'Africa', 'Malawi': 'Africa', 'Mali': 'Africa', 'Mauritania': 'Africa',
    'Mauritius': 'Africa', 'Morocco': 'Africa', 'Mozambique': 'Africa', 'Namibia': 'Africa', 'Niger': 'Africa',
    'Nigeria': 'Africa', 'Rwanda': 'Africa', 'Sao Tome and Principe': 'Africa', 'Senegal': 'Africa',
    'Seychelles': 'Africa', 'Sierra Leone': 'Africa', 'Somalia': 'Africa', 'South Africa': 'Africa',
    'South Sudan': 'Africa', 'Sudan': 'Africa', 'Tanzania': 'Africa', 'Togo': 'Africa', 'Tunisia': 'Africa',
    'Uganda': 'Africa', 'Zambia': 'Africa', 'Zimbabwe': 'Africa',

    # Asia
    'Afghanistan': 'Asia', 'Armenia': 'Asia', 'Azerbaijan': 'Asia', 'Bahrain': 'Asia', 'Bangladesh': 'Asia',
    'Bhutan': 'Asia', 'Brunei': 'Asia', 'Cambodia': 'Asia', 'China': 'Asia', 'Cyprus': 'Asia', 'Georgia': 'Asia',
    'India': 'Asia', 'Indonesia': 'Asia', 'Iran': 'Asia', 'Iraq': 'Asia', 'Israel': 'Asia', 'Japan': 'Asia',
    'Jordan': 'Asia', 'Kazakhstan': 'Asia', 'Kuwait': 'Asia', 'Kyrgyzstan': 'Asia', 'Laos': 'Asia', 'Lebanon': 'Asia',
    'Malaysia': 'Asia', 'Maldives': 'Asia', 'Mongolia': 'Asia', 'Myanmar': 'Asia', 'Nepal': 'Asia', 'North Korea': 'Asia',
    'Oman': 'Asia', 'Pakistan': 'Asia', 'Palestine': 'Asia', 'Philippines': 'Asia', 'Qatar': 'Asia', 'Saudi Arabia': 'Asia',
    'Singapore': 'Asia', 'South Korea': 'Asia', 'Sri Lanka': 'Asia', 'Syria': 'Asia', 'Taiwan': 'Asia', 'Tajikistan': 'Asia',
    'Thailand': 'Asia', 'Timor-Leste': 'Asia', 'Turkey': 'Asia', 'Turkmenistan': 'Asia', 'United Arab Emirates': 'Asia',
    'Uzbekistan': 'Asia', 'Vietnam': 'Asia', 'Yemen': 'Asia',

    # Europe
    'Albania': 'Europe', 'Andorra': 'Europe', 'Austria': 'Europe', 'Belarus': 'Europe', 'Belgium': 'Europe',
    'Bosnia and Herzegovina': 'Europe', 'Bulgaria': 'Europe', 'Croatia': 'Europe', 'Czech Republic': 'Europe',
    'Denmark': 'Europe', 'Estonia': 'Europe', 'Finland': 'Europe', 'France': 'Europe', 'Germany': 'Europe',
    'Greece': 'Europe', 'Hungary': 'Europe', 'Iceland': 'Europe', 'Ireland': 'Europe', 'Italy': 'Europe',
    'Latvia': 'Europe', 'Liechtenstein': 'Europe', 'Lithuania': 'Europe', 'Luxembourg': 'Europe', 'Malta': 'Europe',
    'Moldova': 'Europe', 'Monaco': 'Europe', 'Montenegro': 'Europe', 'Netherlands': 'Europe', 'North Macedonia': 'Europe',
    'Norway': 'Europe', 'Poland': 'Europe', 'Portugal': 'Europe', 'Romania': 'Europe', 'Russia': 'Europe',
    'San Marino': 'Europe', 'Serbia': 'Europe', 'Slovakia': 'Europe', 'Slovenia': 'Europe', 'Spain': 'Europe',
    'Sweden': 'Europe', 'Switzerland': 'Europe', 'Ukraine': 'Europe', 'United Kingdom': 'Europe', 'UK': 'Europe',
    'Vatican City': 'Europe',

    # North America
    'Antigua and Barbuda': 'North America', 'Bahamas': 'North America', 'Barbados': 'North America', 'Belize': 'North America',
    'Canada': 'North America', 'Costa Rica': 'North America', 'Cuba': 'North America', 'Dominica': 'North America',
    'Dominican Republic': 'North America', 'El Salvador': 'North America', 'Grenada': 'North America', 'Guatemala': 'North America',
    'Haiti': 'North America', 'Honduras': 'North America', 'Jamaica': 'North America', 'Mexico': 'North America',
    'Nicaragua': 'North America', 'Panama': 'North America', 'Saint Kitts and Nevis': 'North America',
    'Saint Lucia': 'North America', 'Saint Vincent and the Grenadines': 'North America', 'Trinidad and Tobago': 'North America',
    'United States': 'North America', 'United States (western part)': 'North America', 'United States (eastern part)': 'North America',

    # South America
    'Argentina': 'South America', 'Bolivia': 'South America', 'Brazil': 'South America', 'Chile': 'South America',
    'Colombia': 'South America', 'Ecuador': 'South America', 'Guyana': 'South America', 'Paraguay': 'South America',
    'Peru': 'South America', 'Suriname': 'South America', 'Uruguay': 'South America', 'Venezuela': 'South America',

    # Oceania
    'Australia': 'Oceania', 'Fiji': 'Oceania', 'Kiribati': 'Oceania', 'Marshall Islands': 'Oceania', 'Micronesia': 'Oceania',
    'Nauru': 'Oceania', 'New Zealand': 'Oceania', 'Palau': 'Oceania', 'Papua New Guinea': 'Oceania', 'Samoa': 'Oceania',
    'Solomon Islands': 'Oceania', 'Tonga': 'Oceania', 'Tuvalu': 'Oceania', 'Vanuatu': 'Oceania', 'Cook Islands': 'Oceania',

    # Antarctica
    'Antarctica': 'Antarctica'
}

# Utility Functions
def get_script_name():
    return os.path.splitext(os.path.basename(__file__))[0]

def extract_country_from_url(url):
    match = re.search(r'/country/(.+)\.jpg', url)
    if match:
        return match.group(1).replace('_', ' ')
    return "Unknown"

# Scraping Function
def scrape_transactions(page_num, session, retry_count=0):
    url = URL_TEMPLATE.format(page_num=page_num)
    transactions = []
    try:
        print(f"Scraping page {page_num} (Attempt {retry_count + 1}/{MAX_RETRIES})")
        time.sleep(random.uniform(1, 2))
        response = session.get(url, headers=HEADERS, timeout=30)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')
        transaction_list = soup.select_one('dl.transactions.mt_1')
        if not transaction_list:
            print(f"WARNING: No transaction list found on page {page_num}.")
            return transactions, retry_count + 1
        
        for dd in transaction_list.find_all('dd'):
            sections = dd.find_all('section')
            if len(sections) != 4:
                print(f"WARNING: Transaction on page {page_num} has incorrect number of sections.")
                continue
            buyer_section = sections[0]
            buyer_name = buyer_section.find('span').text.strip() if buyer_section.find('span') else "Unknown"
            country_img = buyer_section.find('img')
            country = extract_country_from_url(country_img['src']) if country_img else "Unknown"
            unit_price = sections[1].text.strip()
            quantity_text = sections[2].text.strip()
            quantity = re.search(r'\d+', quantity_text).group() if re.search(r'\d+', quantity_text) else "0"
            order_date = sections[3].text.strip()
            transactions.append({
                'Name': buyer_name,
                'Country': country,
                'Unit Price': unit_price,
                'Quantity': quantity,
                'Order Date': order_date
            })
        return transactions, retry_count
    except requests.exceptions.RequestException as e:
        print(f"ERROR: Failed to access {url}. Reason: {str(e)}")
        return transactions, retry_count + 1
    except Exception as e:
        print(f"ERROR: Unexpected error processing page {page_num}. Reason: {str(e)}")
        return transactions, retry_count + 1

# Updated Chart Function with improved formatting
def add_bar_chart(worksheet, title, position, cat_col, data_col, min_row, max_row, num_categories):
    chart = BarChart()
    
    # Define data and categories with exact ranges
    data = Reference(worksheet, min_col=data_col, min_row=min_row, max_row=max_row)
    categories = Reference(worksheet, min_col=cat_col, min_row=min_row, max_row=max_row)
    
    # Create a single series explicitly
    series = Series(data, title="Count")
    chart.append(series)
    chart.set_categories(categories)
    
    # Chart properties with improved formatting
    chart.title = title
    chart.width = max(15, min(num_categories * 1.5, 50))  # Auto-expand width based on data
    chart.height = 12
    chart.style = 2
    
    # Enable value axis (left side) with numbers
    chart.y_axis.title = "Count"
    chart.y_axis.majorGridlines = None
    
    # Format category axis (bottom) - rotate labels
    chart.x_axis.title = "Categories"
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.txPr = None  # This will allow text rotation
    
    # Add the chart to worksheet
    worksheet.add_chart(chart, position)
    
    return chart

# Function to auto-adjust column widths with minimum width
def auto_adjust_column_widths(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max((max_length + 2) * 1.2, 10)  # minimum width 10
        worksheet.column_dimensions[column].width = adjusted_width

# Function to add navigation links
def add_navigation_links(worksheet, chart_positions):
    # Freeze first row
    worksheet.freeze_panes = 'A2'
    
    # Add navigation links starting from D2
    col_start = 4  # Column D
    for i, (name, position) in enumerate(chart_positions.items()):
        cell = worksheet.cell(row=1, column=col_start + i)
        cell.value = name
        cell.font = Font(color="0000FF", underline="single")
        cell.alignment = Alignment(horizontal="center")
        
        # Create hyperlink to chart position
        row_num = int(position[1:]) if position[1:].isdigit() else int(re.search(r'\d+', position).group())
        cell.hyperlink = f"#Sheet{worksheet.title[-1]}!A{row_num}"

# Main Function
def main():
    script_name = get_script_name()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, script_name)
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, f"{script_name}.xlsx")
    
    print(f"\nScraper started. Data will be saved in '{script_name}' folder.")
    print("Ensure you have run 'pip install requests beautifulsoup4 pandas openpyxl'.")
    
    all_transactions = []
    with requests.Session() as session:
        for page_num in range(1, TOTAL_PAGES + 1):
            retry_count = 0
            backoff_delay = 5
            while retry_count < MAX_RETRIES:
                transactions, retry_count = scrape_transactions(page_num, session, retry_count)
                if transactions or retry_count >= MAX_RETRIES:
                    all_transactions.extend(transactions)
                    break
                print(f"Retrying page {page_num} after {backoff_delay} seconds...")
                time.sleep(backoff_delay)
                backoff_delay += 1
            if retry_count >= MAX_RETRIES:
                print(f"ERROR: Max retries reached for page {page_num}. Skipping.")
    
    if all_transactions:
        df = pd.DataFrame(all_transactions)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Sheet1: Raw Data
            df.to_excel(writer, sheet_name='Sheet1', startrow=0, index=False)
            worksheet1 = writer.sheets['Sheet1']
            auto_adjust_column_widths(worksheet1)
            
            # Sheet2: Country Analysis
            country_counts = df['Country'].value_counts().sort_values(ascending=False)
            country_counts_df = country_counts.reset_index()
            country_counts_df.columns = ['Country', 'Count']
            country_counts_df.to_excel(writer, sheet_name='Sheet2', startrow=0, index=False)
            worksheet2 = writer.sheets['Sheet2']
            data_min_row = 2  # headers at row 1, data at row 2
            data_max_row = 1 + len(country_counts_df)
            chart_positions = {'Total Country Chart': 'D2'}  # align with header
            add_bar_chart(worksheet2, "Country Occurrences", "D2", 1, 2, data_min_row, data_max_row, len(country_counts_df))
            
            # Sub-charts for continents
            continent_counts = {}
            for country, count in country_counts.items():
                continent = COUNTRY_CONTINENT.get(country, 'Unknown')
                continent_counts[continent] = continent_counts.get(continent, 0) + count
            sorted_continents = sorted(continent_counts.items(), key=lambda x: x[1], reverse=True)
            continent_data = {}
            for country, count in country_counts.items():
                continent = COUNTRY_CONTINENT.get(country, 'Unknown')
                if continent in continent_data:
                    continent_data[continent].append((country, count))
                else:
                    continent_data[continent] = [(country, count)]
            
            seven_continents = ['Africa', 'Asia', 'Europe', 'North America', 'South America', 'Oceania', 'Antarctica']
            current_row = worksheet2.max_row + 20
            for continent in seven_continents:
                if continent in continent_data:
                    continent_df = pd.DataFrame(continent_data[continent], columns=['Country', 'Count'])
                    continent_df = continent_df.sort_values('Count', ascending=False)
                    header_row = current_row
                    continent_df.to_excel(writer, sheet_name='Sheet2', startrow=header_row - 1, index=False)
                    data_min_row = header_row + 1
                    data_max_row = header_row + len(continent_df)
                    chart_position = f"D{header_row}"
                    chart_positions[continent] = chart_position
                    add_bar_chart(worksheet2, f"{continent} Country Occurrences", chart_position, 
                                  1, 2, data_min_row, data_max_row, len(continent_df))
                    current_row = data_max_row + 20
            
            add_navigation_links(worksheet2, chart_positions)
            auto_adjust_column_widths(worksheet2)
            
            # Sheet3: Unit Price Analysis
            unit_price_counts = df['Unit Price'].value_counts().sort_values(ascending=False)
            unit_price_counts_df = pd.DataFrame({
                'Unit Price': unit_price_counts.index,
                'Count': unit_price_counts.values
            })
            unit_price_counts_df.to_excel(writer, sheet_name='Sheet3', startrow=0, index=False)
            worksheet3 = writer.sheets['Sheet3']
            data_min_row = 2
            data_max_row = 1 + len(unit_price_counts_df)
            chart_positions = {'Total Chart': 'D2'}
            add_bar_chart(worksheet3, "Unit Price Occurrences", "D2", 1, 2, data_min_row, data_max_row, len(unit_price_counts_df))
            add_navigation_links(worksheet3, chart_positions)
            auto_adjust_column_widths(worksheet3)
            
            # Sheet4: Quantity Analysis
            quantity_counts = df['Quantity'].value_counts().sort_values(ascending=False)
            quantity_counts_df = pd.DataFrame({
                'Quantity': quantity_counts.index,
                'Count': quantity_counts.values
            })
            quantity_counts_df.to_excel(writer, sheet_name='Sheet4', startrow=0, index=False)
            worksheet4 = writer.sheets['Sheet4']
            data_min_row = 2
            data_max_row = 1 + len(quantity_counts_df)
            chart_positions = {'Total Chart': 'D2'}
            add_bar_chart(worksheet4, "Quantity Occurrences", "D2", 1, 2, data_min_row, data_max_row, len(quantity_counts_df))
            add_navigation_links(worksheet4, chart_positions)
            auto_adjust_column_widths(worksheet4)
            
            # Sheet5: Order Date Analysis
            order_date_counts = df['Order Date'].value_counts().sort_index()
            order_date_counts_df = order_date_counts.reset_index()
            order_date_counts_df.columns = ['Order Date', 'Count']
            order_date_counts_df.to_excel(writer, sheet_name='Sheet5', startrow=0, index=False)
            worksheet5 = writer.sheets['Sheet5']
            data_min_row = 2
            data_max_row = 1 + len(order_date_counts_df)
            chart_positions = {'Total Chart': 'D2'}
            add_bar_chart(worksheet5, "Order Date Occurrences", "D2", 1, 2, data_min_row, data_max_row, len(order_date_counts_df))
            
            # Sub-charts by Year
            year_counts = {}
            for date, count in order_date_counts.items():
                year = date[:4]
                year_counts[year] = year_counts.get(year, 0) + count
            sorted_years = sorted(year_counts.items(), key=lambda x: x[1], reverse=True)
            year_data = {}
            for date, count in order_date_counts.items():
                year = date[:4]
                if year in year_data:
                    year_data[year].append((date, count))
                else:
                    year_data[year] = [(date, count)]
            
            current_row = worksheet5.max_row + 20
            sub_chart_num = 1
            for year, _ in sorted_years:
                if year in year_data:
                    year_df = pd.DataFrame(year_data[year], columns=['Order Date', 'Count'])
                    year_df = year_df.sort_values('Order Date')
                    header_row = current_row
                    year_df.to_excel(writer, sheet_name='Sheet5', startrow=header_row - 1, index=False)
                    data_min_row = header_row + 1
                    data_max_row = header_row + len(year_df)
                    chart_position = f"D{header_row}"
                    chart_positions[f'Sub Chart {sub_chart_num}'] = chart_position
                    add_bar_chart(worksheet5, f"{year} Order Date Occurrences", chart_position, 
                                  1, 2, data_min_row, data_max_row, len(year_df))
                    current_row = data_max_row + 20
                    sub_chart_num += 1
            
            add_navigation_links(worksheet5, chart_positions)
            auto_adjust_column_widths(worksheet5)
            
            # Sheet6: Order Month Analysis
            df['Month'] = df['Order Date'].str[:7]
            month_counts = df['Month'].value_counts().sort_index()
            month_counts_df = month_counts.reset_index()
            month_counts_df.columns = ['Month', 'Count']
            month_counts_df.to_excel(writer, sheet_name='Sheet6', startrow=0, index=False)
            worksheet6 = writer.sheets['Sheet6']
            data_min_row = 2
            data_max_row = 1 + len(month_counts_df)
            chart_positions = {'Total Chart': 'D2'}
            add_bar_chart(worksheet6, "Order Month Occurrences", "D2", 1, 2, data_min_row, data_max_row, len(month_counts_df))
            
            # Sub-charts by Year
            year_month_counts = {}
            for month, count in month_counts.items():
                year = month[:4]
                year_month_counts[year] = year_month_counts.get(year, 0) + count
            sorted_years_month = sorted(year_month_counts.items(), key=lambda x: x[1], reverse=True)
            month_data = {}
            for month, count in month_counts.items():
                year = month[:4]
                if year in month_data:
                    month_data[year].append((month, count))
                else:
                    month_data[year] = [(month, count)]
            
            current_row = worksheet6.max_row + 20
            sub_chart_num = 1
            for year, _ in sorted_years_month:
                if year in month_data:
                    month_df = pd.DataFrame(month_data[year], columns=['Month', 'Count'])
                    month_df = month_df.sort_values('Month')
                    header_row = current_row
                    month_df.to_excel(writer, sheet_name='Sheet6', startrow=header_row - 1, index=False)
                    data_min_row = header_row + 1
                    data_max_row = header_row + len(month_df)
                    chart_position = f"D{header_row}"
                    chart_positions[f'Sub Chart {sub_chart_num}'] = chart_position
                    add_bar_chart(worksheet6, f"{year} Order Month Occurrences", chart_position, 
                                  1, 2, data_min_row, data_max_row, len(month_df))
                    current_row = data_max_row + 20
                    sub_chart_num += 1
            
            add_navigation_links(worksheet6, chart_positions)
            auto_adjust_column_widths(worksheet6)
            
            # Sheet7: Order Year Analysis
            df['Year'] = df['Order Date'].str[:4]
            year_counts = df['Year'].value_counts().sort_index()
            year_counts_df = year_counts.reset_index()
            year_counts_df.columns = ['Year', 'Count']
            year_counts_df.to_excel(writer, sheet_name='Sheet7', startrow=0, index=False)
            worksheet7 = writer.sheets['Sheet7']
            data_min_row = 2
            data_max_row = 1 + len(year_counts_df)
            chart_positions = {'Total Chart': 'D2'}
            add_bar_chart(worksheet7, "Order Year Occurrences", "D2", 1, 2, data_min_row, data_max_row, len(year_counts_df))
            add_navigation_links(worksheet7, chart_positions)
            auto_adjust_column_widths(worksheet7)
        
        print(f"\nScraped {len(all_transactions)} transactions.")
        print(f"Data saved to {excel_path} with improved charts and navigation.")
    else:
        print("\nNo transactions were scraped.")
    
    print(f"\n--- Scraping complete! Data saved in '{script_name}' folder. ---")

if __name__ == "__main__":
    main()